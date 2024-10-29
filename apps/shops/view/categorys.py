import csv
import io

import openpyxl
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.http import HttpResponse
from drf_spectacular.utils import extend_schema
from rest_framework import status
from rest_framework.filters import SearchFilter
from rest_framework.generics import (ListCreateAPIView, RetrieveUpdateDestroyAPIView)
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.views import APIView

from shared.restframework.paginations import PageSortNumberPagination
from shops.models import (Category)
from shops.serializers.category import (CategoryModelSerializer)


@extend_schema(tags=['Category'])
class CategoryCreateAPIView(ListCreateAPIView):
    queryset = Category.objects.all()
    serializer_class = CategoryModelSerializer
    pagination_class = PageSortNumberPagination
    filter_backends = (SearchFilter,)
    search_fields = ('name',)

    def get_queryset(self):
        shop_id = self.kwargs['shop_id']
        return Category.objects.filter(shop_id=shop_id)


@extend_schema(tags=['Category'])
class CategoryUpdateAPIView(RetrieveUpdateDestroyAPIView):
    serializer_class = CategoryModelSerializer
    queryset = Category.objects.all()
    pagination_class = PageSortNumberPagination

    # def category_visibility(request):
    #     category_id = request.data.get('category_id')
    #     show_in_ecommerce = request.data.get('show_in_ecommerce')
    #     category = Category.objects.get(pk=category_id)
    #     category.show_in_ecommerce = show_in_ecommerce
    #     category.save()
    #     return Response({"success": True})


@extend_schema(tags=['CategoryImport'])  # Todo
class CategoryImportAPIView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def get(self, request, *args, **kwargs):
        # Create a sample Excel file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Categories"
        headers = ["ID", "Name", "Parent ID"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        # Save the workbook to a BytesIO object
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        # Create the HttpResponse object with Excel mime type
        response = HttpResponse(output.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=category_template.xlsx'
        return response

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file was uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        # Save the file temporarily
        path = default_storage.save('tmp/' + file.name, ContentFile(file.read()))

        try:
            if file.name.endswith('.csv'):
                with default_storage.open(path, 'r') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        # Process each row here
                        print(row)
                        # You would typically create or update Category objects here
            elif file.name.endswith(('.xls', '.xlsx')):
                workbook = openpyxl.load_workbook(default_storage.open(path))
                sheet = workbook.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Process each row here
                    print(row)
                    # You would typically create or update Category objects here
            else:
                return Response({"error": "Unsupported file format."}, status=status.HTTP_400_BAD_REQUEST)

            return Response({"status": "File uploaded and processed successfully."})
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        finally:
            # Clean up the temp file
            default_storage.delete(path)
